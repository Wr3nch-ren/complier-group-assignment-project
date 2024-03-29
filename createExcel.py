import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (Alignment, Font)
import openpyxl

file = pd.read_csv("output.csv", header=None)
output = 'result.xlsx'

workbook = openpyxl.Workbook()
workbook.save(output)
workbook.close()

entries = []
for i in file[1] :
    if type(i) == float : 
        continue
    if len(i) > 1:
        continue
    if i not in entries :
        entries.append(i)

accept = []
states = []
for i in range(len(file[0])) :
    if len(file[0][i]) > 1:
        continue
    if 'q'+file[0][i] not in states :
        states.append('q'+ file[0][i])
        print(file[3][i])
        if file[3][i]:
            accept.append('q'+file[0][i])

transition = []
state = file[0][0]
lst = []
for i in range(len(file[2])) :
    if file[0][i] == state :
        if file[2][i] == 0 :
            lst.append("∅")
        else:
            lst.append(f"q{file[2][i]:.0f}")
    else :
        transition.append(lst[:])
        state = file[0][i]
        lst = []
        if file[2][i] == 0 :
            lst.append("∅")
        else:
            lst.append(f"q{file[2][i]:.0f}")


start = states[0]

df = pd.DataFrame(transition, index=states, columns=entries)
df.to_excel(output, index=True, header=True)

workbook = load_workbook(filename=output)
sheet = workbook["Sheet1"]

sheet["A1"] = "DFA"
sheet["A1"].alignment = Alignment(horizontal='center')
sheet["A1"].font = Font(bold=True)

sheet.append((" "," "))
sheet.append(("Q :",f"{{{','.join(states)}}}"))
sheet.append(("Σ : ",f"{{{','.join(entries)}}}"))
sheet.append(("Start : ",start))
sheet.append(("Accept : ",f"{{{','.join(accept)}}}"))

print(df)
workbook.save(filename=output)